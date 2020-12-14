import pandas
import urllib
import urllib.request
import os
import openpyxl

#pandas.read_json("input.json",lines=True).to_excel("output2.xlsx")
filename="d:\output - 副本.xlsx"
inwb = openpyxl.load_workbook(filename)
sheetnames = inwb.get_sheet_names()
ws = inwb.get_sheet_by_name(sheetnames[1])
rows = ws.max_row
print(rows)
#cols = ws.max_column
for r in range(1,rows):
	value = ws.cell(r,3).value
	if not value is None:
		if not os.path.exists("d:\\file\\"+ws.cell(r,4).value+"\\"+ws.cell(r,2).value):
			os.makedirs("d:\\file\\"+ws.cell(r,4).value+"\\"+ws.cell(r,2).value)
		if not os.path.exists("d:\\file\\"+ws.cell(r,4).value+"\\"+ws.cell(r,2).value+"\\"+value[43:100]):
			print(r+"-"+ws.cell(r,4).value)
			urllib.request.urlretrieve(ws.cell(r,3).value, "d:\\file\\"+ws.cell(r,4).value+"\\"+ws.cell(r,2).value+"\\"+value[43:100])